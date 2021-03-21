def empty_result_xlsx_files(path_to_file):
    #path_to_file="/media/alex/新加卷/OBS_Recordings_T/Data_science/Python/My_python_programs/IOYU/http_to_https"
    import os
    import openpyxl
    file_names=['http_to_https_results.xlsx','http_to_https_results_with_https_sheet.xlsx']
    #from openpyxl import Workbook
    os.chdir ( path_to_file )
    #print ( path_to_file )
    for file_name in file_names:
        #print ( os.path.join ( path_to_file , file_name ) )
        if os.path.exists ( os.path.join ( path_to_file , file_name ) ):
            print(f'file {file_name} exists. Now I am gonna clean it up for future notes')
            wb=openpyxl.load_workbook(os.path.join ( path_to_file , file_name ))
            list_of_sheets=wb.sheetnames
            #print(list_of_sheets)

            for sheet in list_of_sheets:
                #print ('sheet is',sheet)
                cs=wb[sheet]
                cs.delete_rows(1,cs.max_row)
                cs.delete_cols ( 1 , cs.max_column )
                #print ( 'cs=' , cs )
                wb.remove(cs)
                #print ('cs is removed')
            wb.create_sheet('Sheet')
            wb.save ( os.path.join ( path_to_file , file_name ) )



        else:
            print(f'file {file_name} does not exist. There is nothing to clean up')
        #wb.create_sheet('List_of_http_and_links')


#empty_result_xlsx_files()