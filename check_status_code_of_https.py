
def check_status_code_of_https(path_to_file):
    ##
    import openpyxl
    import requests
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill
    import os
    #from Change_http_into_https import path_to_file
    #global path_to_file
    wb=openpyxl.load_workbook ( os.path.join (path_to_file, 'http_to_https_results.xlsx' ) )
    #wb = openpyxl.load_workbook(
    #    "/home/alex/OBS_recordings/Link_to_OBS_Recordings_T/Data_science/Python/My_python_programs/IOYU/http_to_https/http_to_https_results_test.xlsx", read_only = True)
    ws = wb.active
    #print(ws.title)
    #wb.create_sheet('https_addresses')
    wb2=Workbook()
    ws2_1=wb2.create_sheet('link to page, http, code',0)
    ws2_2=wb2.create_sheet('link to page, https, code',1)
    # for row_counter, row_value in enumerate(inner_list_of_tuple_of_link_and_code):
    #     for column_counter, cell_value in enumerate(inner_list_of_tuple_of_link_and_code[row_counter]):
    #         #print (inner_list_of_tuple_of_link_and_code[row_counter][column_counter])
    #         ws.cell(row=row_counter+1, column=column_counter+1).value=inner_list_of_tuple_of_link_and_code[row_counter][column_counter]
    ##
    ##
    a=ws.iter_rows(values_only = True)
    list_of_links=list(a)
    #print(*list_of_links, sep = '\n')
    #print ("link=", *list_of_links, sep="\n")
    list_of_https_addresses_and_status_codes=[]
    for row_counter, value in enumerate(list_of_links):
        https_address=list_of_links[row_counter][1].replace("http://173904.selcdn.com" , "https://173904.selcdn.ru")
        link_to_page=list_of_links[row_counter][0]
        print("link_to_page=", link_to_page)
        response = requests.head ( https_address )
        code = response.status_code
        print ('https_address_is', https_address, 'code=', code)
        list_of_https_addresses_and_status_codes.append([link_to_page,https_address,code])
    ##
    #print(*list_of_https_addresses_and_status_codes, sep='\n')
    #wb.active = wb.sheetnames.index ( 'https_addresses' )
    #ws = wb.active
    #print ( ws.title )
    ##
    ws2_1=wb2['link to page, http, code']

    header1=['on this page a buggy http occurred','a link with http and selcdn','status code']
    ws2_1.append(header1)
    ws2_1['A1'].fill=PatternFill(fgColor = 'ffff40', fill_type = 'solid')
    ws2_1['B1'].fill = PatternFill ( fgColor = 'ffff40' , fill_type = 'solid' )
    ws2_1['C1'].fill = PatternFill ( fgColor = 'ffff40' , fill_type = 'solid' )
    #print ('list_of_links=',*list_of_links, sep = '\n')
    for row in list_of_links:
        ws2_1.append(row)
    ws2_2 = wb2['link to page, https, code']
    #print ( 'list_of_https_addresses_and_status_codes=' , *list_of_https_addresses_and_status_codes, sep = '\n' )
    header2=['on this page a buggy http occurred','http link with s added and com to ru changed','status_code (200 is good)']
    ws2_2.append ( header2 )
    ws2_2['A1'].fill = PatternFill ( fgColor = 'ffff40' , fill_type = 'solid' )
    ws2_2['B1'].fill = PatternFill ( fgColor = 'ffff40' , fill_type = 'solid' )
    ws2_2['C1'].fill = PatternFill ( fgColor = 'ffff40' , fill_type = 'solid' )
    for row in list_of_https_addresses_and_status_codes:
        ws2_2.append(row)
    wb2.save(os.path.join (path_to_file, 'http_to_https_results_with_https_sheet.xlsx' ))
    ##
#check_status_code_of_https()



##

